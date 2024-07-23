import openpyxl
from collections import defaultdict
import io
import os
import xlrd
import datetime as dt
from django.conf import settings
from .models import MyFiles, Faculty, Qualification, Group, Weekday, Discipline, LessonType, \
    Department, Position, Teacher, Classroom, LessonTime, Subgroup, Lesson, ExamTime, Exam
import xml.etree.ElementTree as ET


# форматирование даты, если значение ячейки является числом с плавающей точкой
def get_formatted_date(cell_value, workbook):
    if isinstance(cell_value, float):
        date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
        return dt.datetime(*date_tuple[:3]).strftime('%d.%m.%Y')
    return str(cell_value)


# обработка ячейки с датой, независимо от того, является ли значение числом с плавающей точкой или нет
def process_date_cell(cell_value, workbook):
    if isinstance(cell_value, float):
        return get_formatted_date(cell_value, workbook)
    return str(cell_value)


# форматирование времени в одно значение
def format_time(hour_value, min_value):
    hour_value = int(float(hour_value)) if hour_value else 0
    min_value = int((float(min_value) % 1) * 100) if min_value else 0
    return f"{hour_value}:{min_value:02d}"


# обработка пустых ячеек
def handle_empty_cell(header, default_columns):
    if header in default_columns:
        return 'null'  # значение по умолчанию
    else:
        return 'empty'  # ошибочное значение


# перевод float в int
def convert_float_to_int(columns):
    return [int(float(x)) if isinstance(x, (float, str)) and str(x).replace('.', '', 1).isdigit() else x for x in columns]


# ########################################### ЗАНЯТИЯ ЗАНЯТИЯ ЗАНЯТИЯ
def parser_class_rasp_xls(file_path, filename, folder, file_type):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_name('bd')
    headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

    header_to_xml_map = {
        'семестр': 'semester_index',
        'Фак-т': 'faculty',
        'Курс': 'course',
        'Группа:': 'group',
        'Дисциплина': 'discipline',
        'каф.': 'department',
        'Преподаватель': 'teacher',
        '№ груп.': 'num_group',
        'кол-во человек': 'num_students',
        'Вид занятия': 'lesson_type',
        'аудитория': 'auditorium',
        'Корпус': 'building',
        'Дата': 'start_date',
        'Дата\n(конец\nдиапазона)': 'end_date',
        'День': 'day_of_week',
        'Нач.з.час': 'start_time',
        'Кон.з.час': 'end_time'
    }

    default_columns = ['№ груп.', 'кол-во человек']
    date_columns = ['Дата', 'Дата\n(конец\nдиапазона)']
    columns_to_convert = ['Курс', 'каф.', '№ груп.', 'кол-во человек', 'аудитория']

    root = ET.Element('lessons')
    file_type_element = ET.Element('file_type')
    file_type_element.text = file_type
    root.insert(0, file_type_element)

    for row in range(1, sheet.nrows):
        lesson = ET.SubElement(root, 'lesson')
        for header, xml_element in header_to_xml_map.items():
            index = headers.index(header)
            cell_value = sheet.cell_value(row, index)

            if cell_value == '':
                cell_value = handle_empty_cell(header, default_columns)

            if header in date_columns:
                cell_value = process_date_cell(cell_value, workbook)

            elif header in ['Нач.з.час', 'Кон.з.час']:
                hour_value = cell_value
                min_value = sheet.cell_value(row, index + 1)
                cell_value = format_time(hour_value, min_value)

            elif header in columns_to_convert:
                cell_value = convert_float_to_int([cell_value])[0]

            elif header == 'Преподаватель':
                teacher_value = sheet.cell_value(row, index)
                if teacher_value == '':
                    ET.SubElement(lesson, 'position').text = 'null'
                    ET.SubElement(lesson, 'teacher').text = ''
                else:
                    parts = teacher_value.split(' ', 1)
                    position = parts[0]
                    name = parts[1]
                    ET.SubElement(lesson, 'position').text = position
                    ET.SubElement(lesson, 'teacher').text = name
                continue

            elif header == 'Дисциплина':
                discipline_value = sheet.cell_value(row, index)
                if discipline_value == '':
                    ET.SubElement(lesson, 'week').text = 'null'
                    ET.SubElement(lesson, 'discipline').text = ''
                else:
                    if any(char in discipline_value for char in ['I', 'II']):
                        week, discipline = discipline_value.split(' ', 1)
                        ET.SubElement(lesson, 'week').text = week
                        ET.SubElement(lesson, 'discipline').text = discipline
                    else:
                        ET.SubElement(lesson, 'week').text = 'null'
                        ET.SubElement(lesson, 'discipline').text = discipline_value
                continue

            ET.SubElement(lesson, xml_element).text = str(cell_value)

    file_name, file_extension = filename.rsplit('.', 1)
    output_folder = os.path.join(settings.MEDIA_ROOT, 'create_xml', folder)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    output_xml_file = os.path.join(output_folder, f'{file_name}.xml')
    tree = ET.ElementTree(root)
    tree.write(output_xml_file)



    tree = ET.parse(output_xml_file)
    root = tree.getroot()

    # Создаем новый корневой элемент
    new_root = ET.Element('lessons')
    file_type_element = ET.Element('file_type')
    file_type_element.text = file_type
    new_root.append(file_type_element)

    # Создаем словарь для хранения групп и их студентов
    groups = {}

    # Проходим по всем урокам
    for lesson in root.findall('.//lesson'):
        # Получаем информацию о группе
        group_name = lesson.find('group').text
        num_students = lesson.find('num_students').text
        if num_students == 'null':
            # Если количество студентов не указано, ищем в других уроках
            for other_lesson in root.findall('.//lesson'):
                if other_lesson.find('group').text == group_name and other_lesson.find('num_students').text != 'null':
                    num_students = other_lesson.find('num_students').text
                    break
        groups[group_name] = num_students

        # Создаем новый элемент semester
        semester_name = lesson.find('semester_index').text
        semester = new_root.find(f"semester[@name='{semester_name}']")
        if semester is None:
            semester = ET.SubElement(new_root, 'semester', attrib={'name': semester_name})

        # Создаем новый элемент faculty
        faculty_name = lesson.find('faculty').text
        faculty = semester.find(f"faculty[@name='{faculty_name}']")
        if faculty is None:
            faculty = ET.SubElement(semester, 'faculty', attrib={'name': faculty_name})

        # Создаем новый элемент course
        course_name = lesson.find('course').text
        course = faculty.find(f"course[@name='{course_name}']")
        if course is None:
            course = ET.SubElement(faculty, 'course', attrib={'name': course_name})

        # Создаем новый элемент group
        group = course.find(f"group[@name='{group_name}']")
        if group is None:
            group = ET.SubElement(course, 'group', attrib={'name': group_name})
            group.append(ET.Element('num_students', text=num_students))

        # Создаем новый элемент week
        week_name = lesson.find('week').text
        week = group.find(f"week[@name='{week_name}']")
        if week is None:
            week = ET.SubElement(group, 'week', attrib={'name': week_name})

        # Создаем новый элемент day
        day_name = lesson.find('day_of_week').text
        day = week.find(f"day[@name='{day_name}']")
        if day is None:
            day = ET.SubElement(week, 'day', attrib={'name': day_name})

        # Создаем новый элемент lesson
        lesson_element = ET.SubElement(day, 'lesson')
        for elem in lesson:
            if elem.tag not in ['semester_index', 'faculty', 'course', 'group', 'week', 'day_of_week', 'num_students']:
                lesson_element.append(elem)

    # Создание нового XML файла
    new_tree = ET.ElementTree(new_root)
    new_tree.write(output_xml_file)

    return output_folder, folder


# ########################################### ЭКЗАМЕНЫ ЭКЗАМЕНЫ ЭКЗАМЕНЫ
def parser_exam_rasp(file_path, filename, folder, file_type):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_name('bd')
    headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

    header_to_xml_map = {
        'семестр': 'semester_index',
        'Фак-т': 'faculty',
        'Курс': 'course',
        'Группа:': 'group',
        'Дисциплина': 'discipline',
        'каф.': 'department',
        'Преподаватель': 'teacher',
        'кол-во человек': 'num_students',
        'Вид занятия': 'lesson_type',
        'Корпус': 'building',
        'Дата': 'start_date',
        'День': 'day_of_week',
        'Нач.з.час': 'start_time'
    }

    default_columns = []
    date_columns = ['Дата']
    columns_to_convert = ['Курс', 'каф.', 'кол-во человек', 'аудитория']

    root = ET.Element('exams')
    file_type_element = ET.Element('file_type')
    file_type_element.text = file_type
    root.insert(0, file_type_element)

    for row in range(1, sheet.nrows):
        exam = ET.SubElement(root, 'exam')
        for header, xml_element in header_to_xml_map.items():
            index = headers.index(header)
            cell_value = sheet.cell_value(row, index)

            if cell_value == '':
                cell_value = handle_empty_cell(header, default_columns)

            if header in date_columns:
                cell_value = process_date_cell(cell_value, workbook)

            elif header in ['Нач.з.час']:
                hour_value = cell_value
                min_value = sheet.cell_value(row, index + 1)
                cell_value = format_time(hour_value, min_value)

            elif header in columns_to_convert:
                cell_value = convert_float_to_int([cell_value])[0]

            elif header == 'Преподаватель':
                teacher_value = sheet.cell_value(row, index)
                if teacher_value == '':
                    ET.SubElement(exam, 'position').text = 'null'
                    ET.SubElement(exam, 'teacher').text = ''
                else:
                    parts = teacher_value.split(' ', 1)
                    position = parts[0]
                    name = parts[1]
                    ET.SubElement(exam, 'position').text = position
                    ET.SubElement(exam, 'teacher').text = name
                continue

            ET.SubElement(exam, xml_element).text = str(cell_value)

    file_name, file_extension = filename.rsplit('.', 1)
    output_folder = os.path.join(settings.MEDIA_ROOT, 'create_xml', folder)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    output_xml_file = os.path.join(output_folder, f'{file_name}.xml')
    tree = ET.ElementTree(root)
    tree.write(output_xml_file)
    return output_folder, folder


# ########################################### СВЕДЕНИЯ СВЕДЕНИЯ СВЕДЕНИЯ
def parser_sved(file_path, folder, file_type):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = [sheet.cell(row=8, column=col).value for col in range(1, sheet.max_column + 1)]

    header_to_xml_map = {
        'Наименование дисциплин\nпо учебному плану': 'discipline',
        'ФИО\nлектора': 'teacher_lect',
        'Количество часов в неделю лекций': 'hour_week_lect',
        'Шифр\nгруппы лекции': 'group_lect',
        'ФИО, руководителя\nпрактических занятий\n(без разделения\nна подгруппы)': 'teacher_pract',
        'Количество часов в неделю практик': 'hour_week_pract',
        'Шифр\nгруппы практики': 'group_pract',
    }

    # Создание XML документа
    output_folder = os.path.join(settings.MEDIA_ROOT, 'create_xml', folder)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    output_xml_file = os.path.join(output_folder, 'sved.xml')

    root = ET.Element('data')
    file_type_element = ET.Element('file_type')
    file_type_element.text = file_type
    root.insert(0, file_type_element)

    for row in range(9, sheet.max_row + 1):
        has_discipline_none = False
        has_teacher_pract_none = False

        for col, header in enumerate(headers, start=1):
            cell_value = sheet.cell(row=row, column=col).value
            if header == 'Наименование дисциплин\nпо учебному плану':
                if cell_value is None or cell_value == 'None':
                    has_discipline_none = True
            elif header == 'ФИО, руководителя\nпрактических занятий\n(без разделения\nна подгруппы)':
                if cell_value is None or cell_value == 'None':
                    has_teacher_pract_none = True

        if not (has_discipline_none and has_teacher_pract_none):
            data_row = ET.SubElement(root, 'row')
            for col, header in enumerate(headers, start=1):
                cell_value = sheet.cell(row=row, column=col).value
                if header in header_to_xml_map:
                    xml_header = header_to_xml_map[header]
                    ET.SubElement(data_row, xml_header).text = str(cell_value)

    tree = ET.ElementTree(root)
    tree.write(output_xml_file)

    # Парсинг созданного XML файла
    tree = ET.parse(output_xml_file)
    root = tree.getroot()
    previous_values = {}

    for row in root.findall('row'):
        discipline = row.find('discipline').text
        teacher_lect = row.find('teacher_lect').text
        hour_week_lect = row.find('hour_week_lect').text
        group_lect = row.find('group_lect').text

        if discipline == 'None' and teacher_lect == 'None' and hour_week_lect == 'None' and group_lect == 'None':
            row.find('discipline').text = previous_values.get('discipline', '')
            row.find('teacher_lect').text = previous_values.get('teacher_lect', '')
            row.find('hour_week_lect').text = previous_values.get('hour_week_lect', '')
            row.find('group_lect').text = previous_values.get('group_lect', '')

        previous_values = {
            'discipline': discipline,
            'teacher_lect': teacher_lect,
            'hour_week_lect': hour_week_lect,
            'group_lect': group_lect
        }

    for row in root.findall('row'):
        discipline = row.find('discipline')

        # Изменение значения в <discipline> в соответствии с требованиями
        if discipline.text and "[выбранные дисциплины]" in discipline.text:
            discipline.text = discipline.text[:discipline.text.find("[выбранные дисциплины]")].rstrip() + " [выбранные дисциплины]"

    for row in root.findall('row'):
        for tag in ['teacher_lect', 'teacher_pract']:
            teacher = row.find(tag)
            if teacher is not None and teacher.text != 'None':
                teacher_words = teacher.text.split()
                if len(teacher_words) > 3:
                    position = ' '.join(teacher_words[:-3])
                    teacher.text = ' '.join(teacher_words[-3:])
                else:
                    position = ''

                position_element = ET.SubElement(row, f"position_{tag.split('_')[1]}")
                position_element.text = position

    tree.write(output_xml_file)


# ########################################### НАГРУЗКА НАГРУЗКА НАГРУЗКА
def parser_workload(file_path, folder, file_type, file_semester):
    workbook = openpyxl.load_workbook(file_path)

    # Создание XML документа
    output_folder = os.path.join(settings.MEDIA_ROOT, 'create_xml', folder)
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    output_xml_file = os.path.join(output_folder, 'workload.xml')

    root = ET.Element('data')
    file_type_element = ET.Element('file_type')
    file_type_element.text = file_type
    root.insert(0, file_type_element)

    # Проверяем, что имя состоит из двух слов (Фамилия и инициалы)
    for sheet_name in workbook.sheetnames:
        if sheet_name.count(' ') == 1 and 'Вакансия' not in sheet_name:  # Проверяем условия на формат имени листа
            last_name, first_middle_name = sheet_name.split(' ')

            if last_name and first_middle_name:
                sheet = workbook[sheet_name]
                cell_value = sheet['B8'].value

                if cell_value:
                    # Разделяем строку по пробелам и берем последние три слова
                    words = cell_value.split()
                    if len(words) >= 3:
                        full_name = ' '.join(words[-3:])

                        # Проверяем наличие текста "О С Е Н Н И Й    С Е М Е С Т Р" в столбце A
                        semester = None
                        search_text = None

                        if file_semester == 'autumn':
                            search_text = "О С Е Н Н И Й    С Е М Е С Т Р"
                        elif file_semester == 'spring':
                            search_text = "В Е С Е Н Н И Й    С Е М Е С Т Р"

                        for idx, row in enumerate(
                                sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True),
                                start=1):
                            if search_text in str(row[0]):
                                semester = 'осенний' if file_semester == 'autumn' else 'весенний'
                                start_row = idx + 2  # Переходим на 2 ячейки вниз

                                teacher_element = ET.Element('teacher')
                                full_name_element = ET.Element('full_name')
                                semester_element = ET.Element('semester')
                                full_name_element.text = full_name
                                semester_element.text = semester
                                teacher_element.append(full_name_element)
                                teacher_element.append(semester_element)
                                disciplines_element = ET.Element('disciplines')

                                for cell_idx, cell in enumerate(
                                        sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True)):
                                    if cell[0] == "Всего за семестр":
                                        break
                                    if cell[0] not in ["Всего по основной нагрузке",
                                                       "Всего по дополнительной нагрузке"]:
                                        hour_lect_value = sheet.cell(row=cell_idx + start_row, column=5).value
                                        hour_pract_value = sheet.cell(row=cell_idx + start_row, column=6).value

                                        if hour_lect_value is not None or hour_pract_value is not None:
                                            discipline_text = str(cell[0])
                                            if "[выбранные дисциплины]" in discipline_text:
                                                discipline_text = discipline_text.split("[выбранные дисциплины]")[
                                                                      0].strip() + " [выбранные дисциплины]"

                                            discipline_element = ET.Element('discipline')
                                            discipline_element.text = discipline_text

                                            group_element = ET.Element('group')
                                            group_element.text = str(
                                                sheet.cell(row=cell_idx + start_row, column=2).value)
                                            discipline_element.append(group_element)

                                            hour_lect_element = ET.Element('hour_lect')
                                            hour_lect_element.text = str(hour_lect_value)
                                            discipline_element.append(hour_lect_element)

                                            hour_pract_element = ET.Element('hour_pract')
                                            hour_pract_element.text = str(hour_pract_value)
                                            discipline_element.append(hour_pract_element)

                                            disciplines_element.append(discipline_element)
                                teacher_element.append(disciplines_element)
                                root.append(teacher_element)

    tree = ET.ElementTree(root)
    tree.write(output_xml_file)


# Сохранение данных (в том числе итог проверки) об отчете в бд
def xml_to_bd(folder_path, name, folder):
    year, semester = folder.split('_')

    # Чтение из файла
    file_tree = ET.parse(os.path.join(folder_path, name))
    file_root = file_tree.getroot()
    report_element = file_root.find("report")

    if report_element is not None:
        file_type = report_element.text
    else:
        file_type = "N/A"

    # Поиск или создание записи в базе данных
    existing_file = MyFiles.objects.filter(file=name, year=year, semester=semester, directory_path=folder_path).first()
    if existing_file:
        existing_file.file_type = file_type
        existing_file.save()
        return f"Запись в базе данных обновлена: {existing_file}"
    else:
        new_file = MyFiles(file=name, file_type=file_type, year=year, semester=semester, directory_path=folder_path)
        new_file.save()
        return f"Новая запись создана в базе данных: {new_file}"


# ########################################### Проверка расписания занятий
def check_xml_class(output_folder, folder):
    folder_path = output_folder
    print('папка расписания для проверки', folder_path)
    folder = folder
    # Выбираются файлы, кроме отчетов
    files_list = [f for f in os.listdir(folder_path) if f not in ['report_class_rasp.xml', 'report_exam_rasp.xml']]

    class_rasp_file = None
    workload_file = None
    sved_file = None

    teacher_disciplines_errors_str = ''
    group_disciplines_errors_str = ''
    disciplines_hour_lect_errors_str = ''
    hour_pract_str = ''

    for file in files_list:
        tree = ET.parse(os.path.join(folder_path, file))
        root = tree.getroot()
        file_type = root.find('.//file_type')
        if file_type is not None:
            if file_type.text == 'class_rasp':
                class_rasp_file = file
            elif file_type.text == 'workload':
                workload_file = file
            elif file_type.text == 'sved':
                sved_file = file

    tree = ET.parse(os.path.join(folder_path, class_rasp_file))
    root = tree.getroot()

    # Словарь для хранения дисциплин по группам
    group_disciplines = {}

    # Обход элементов XML для групп
    for group in root.iter('group'):
        group_name = group.attrib.get('name')
        disciplines = set()

        # Находим все элементы <discipline> внутри текущей группы и сохраняем их
        for discipline_elem in group.iter('discipline'):
            discipline = discipline_elem.text
            disciplines.add(discipline)

        # Добавляем список дисциплин для текущей группы в словарь
        group_disciplines[group_name] = disciplines

    # Словарь для хранения дисциплин по преподавателям
    teacher_disciplines = {}

    # Обход элементов XML для преподавателей
    for lesson in root.iter('lesson'):
        teacher = lesson.find('teacher').text
        discipline = lesson.find('discipline').text

        if teacher in teacher_disciplines:
            teacher_disciplines[teacher].add(discipline)
        else:
            teacher_disciplines[teacher] = {discipline}

    # Словарь для хранения данных о лекциях
    data_dict = {}

    for group in root.findall('.//group'):
        group_name = group.attrib.get('name')
        for week in group.findall('.//week'):
            week_name = week.attrib.get('name')
            for lesson in week.findall('.//lesson'):
                lesson_type = lesson.find('.lesson_type').text
                if lesson_type == "лекция":
                    discipline = lesson.find('.discipline').text
                    teacher = lesson.find('.teacher').text

                    data_dict.setdefault((group_name, week_name), []).append({
                        'discipline': discipline,
                        'teacher': teacher,
                        'group': group_name,
                        'week': week_name
                    })

    #for (group, week), lessons in data_dict.items():
    #    for lesson in lessons:
    #        print(f'Группа: {lesson["group"]}, Неделя: {lesson["week"]}')
    #        print(f'Дисциплина: {lesson["discipline"]}')
    #        print(f'Преподаватель: {lesson["teacher"]}')
    #        print('---')

    new_data_dict = {}

    for key, lessons in data_dict.items():
        for lesson in lessons:
            group = lesson['group']
            discipline = lesson['discipline']
            teacher = lesson['teacher']
            week = key[1]
            hours = 1 if week in ['I', 'II'] else 2

            new_data_dict.setdefault((group, discipline, teacher), 0)
            new_data_dict[(group, discipline, teacher)] += hours

    final_dict = {}
    for (group, discipline, teacher), total_hours in new_data_dict.items():
        final_dict.setdefault(group, []).append({
            'discipline': discipline,
            'teacher': teacher,
            'total_hours': total_hours
        })

    #for group, lessons in final_dict.items():
    #    for lesson in lessons:
    #        print(f'Группа: {group}')
    #        print(f'Дисциплина: {lesson["discipline"]}')
    #        print(f'Преподаватель: {lesson["teacher"]}')
    #        print(f'Общее количество часов в неделю: {lesson["total_hours"]}')
    #        print('---')

    # Словарь для хранения данных о практиках
    data_pract = {}

    for semester in root.findall('.//semester'):
        for faculty in semester.findall('.//faculty'):
            for course in faculty.findall('.//course'):
                for group in course.findall('.//group'):
                    group_name = group.get('name')
                    if group_name not in data_pract:
                        data_pract[group_name] = {}  # создаем пустой словарь для группы

                    for week in group.findall('.//week'):
                        week_name = week.get('name')
                        for lesson in week.findall('.//lesson'):
                            lesson_type = lesson.find('lesson_type').text
                            if lesson_type == 'практика':
                                discipline = lesson.find('discipline').text
                                teacher = lesson.find('teacher').text.split()[0]

                                if discipline not in data_pract[group_name]:
                                    data_pract[group_name][discipline] = {}

                                if teacher not in data_pract[group_name][discipline]:
                                    data_pract[group_name][discipline][teacher] = []

                                data_pract[group_name][discipline][teacher].append(week_name)

    data_pract_new = {}
    for group, disciplines in data_pract.items():
        data_pract_new[group] = {}
        for discipline, teachers in disciplines.items():
            data_pract_new[group][discipline] = {}
            for teacher, weeks in teachers.items():
                hours = 0
                for week in weeks:
                    if week == 'I' or week == 'II':
                        hours += 1
                    elif week == 'null':
                        hours += 2
                data_pract_new[group][discipline][teacher] = hours


    # Проверка того, что преподаватель одновременно ведет только одну пару
    lessons_by_teacher = defaultdict(list)
    # Парсинг XML-файла и заполнение данных о занятиях
    for semester in root.iter('semester'):
        for faculty in semester.iter('faculty'):
            for course in faculty.iter('course'):
                for group in course.iter('group'):
                    group_name = group.get('name')
                    for week in group.iter('week'):
                        for day in week.iter('day'):
                            for lesson in day.iter('lesson'):
                                teacher = lesson.find('teacher').text
                                week_name = week.get('name')
                                day_name = day.get('name')
                                start_time = lesson.find('start_time').text
                                end_time = lesson.find('end_time').text
                                discipline = lesson.find('discipline').text
                                lesson_type = lesson.find('lesson_type').text
                                auditorium = lesson.find('auditorium').text
                                # Сохраняем занятия по преподавателям
                                lessons_by_teacher[teacher].append({
                                    'group': group_name,
                                    'week': week_name,
                                    'day': day_name,
                                    'start_time': start_time,
                                    'end_time': end_time,
                                    'discipline': discipline,
                                    'lesson_type': lesson_type,
                                    'auditorium': auditorium,
                                })
    # Проверка на конфликты
    conflicts = defaultdict(list)
    for teacher, lessons in lessons_by_teacher.items():
        # Словарь для хранения временных интервалов, где ключом будут остальные параметры занятий
        temp_dict = defaultdict(list)
        for lesson in lessons:
            key = (lesson['group'], lesson['week'], lesson['day'], lesson['start_time'], lesson['end_time'])
            temp_dict[key].append(lesson)

        # Проверка на конфликты
        for key, conflict_lessons in temp_dict.items():
            if len(conflict_lessons) > 1:
                conflicts[teacher].append(conflict_lessons)

    # Сохранение вывода в переменную
    output = io.StringIO()
    if not conflicts:
        output.write("Преподаватели ведут одно занятие: Успешно\n")

    for teacher, conflict_lessons_groups in conflicts.items():
        output.write(f"Конфликт занятий у преподавателя {teacher}:\n")
        for conflict_lessons in conflict_lessons_groups:
            for lesson in conflict_lessons:
                output.write(
                    f"  Занятие: группа {lesson['group']}, дисциплина {lesson['discipline']} ({lesson['lesson_type']}), {lesson['auditorium']} ауд., неделя {lesson['week']}, день {lesson['day']}, время {lesson['start_time']}-{lesson['end_time']}\n")
            output.write('-' * 50 + '\n')  # Разделитель для отдельных групп конфликтов

    # Получение строки
    teacher_conf_str = output.getvalue()
    output.close()

    # Теперь строку можно использовать как угодно
    print(teacher_conf_str)





    # Проверяем, что workload_file не равен None
    if workload_file is not None:
        tree_workload = ET.parse(os.path.join(folder_path, workload_file))
        root_workload = tree_workload.getroot()

        # Словарь для хранения дисциплин по преподавателям из workload_file
        teacher_disciplines_workload = {}

        # Обход элементов XML для преподавателя
        for teacher_elem in root_workload.iter('teacher'):
            teacher_name = teacher_elem.find('full_name').text
            disciplines = set()

            # Находим все элементы <discipline> внутри текущего преподавателя
            for discipline_elem in teacher_elem.iter('discipline'):
                discipline_name = discipline_elem.text.strip().split('\n')[0]

                # Удаляем "[выбранные дисциплины]" из названия дисциплины, если она присутствует
                if "[выбранные дисциплины]" in discipline_name:
                    discipline_name = discipline_name.replace("[выбранные дисциплины]", "").strip()

                disciplines.add(discipline_name)

            # Добавляем список дисциплин для текущего преподавателя в словарь
            teacher_disciplines_workload[teacher_name] = disciplines

        # Сравнение дисциплин каждого преподавателя
        teacher_disciplines_errors = []
        teacher_disciplines_success = True

        for teacher, disciplines in teacher_disciplines.items():
            last_name = teacher.split()[0]

            found = False

            for teacher_workload, disciplines_workload in teacher_disciplines_workload.items():
                last_name_workload = teacher_workload.split()[0]

                if last_name == last_name_workload:
                    missing_disciplines = disciplines_workload - disciplines
                    extra_disciplines = disciplines - disciplines_workload

                    if missing_disciplines:
                        error = f"У преподавателя {teacher} не хватает дисциплин: {', '.join(missing_disciplines)}\n"
                        teacher_disciplines_errors.append(error)
                        teacher_disciplines_success = False
                    if extra_disciplines:
                        error = f"У преподавателя {teacher} лишняя дисциплина: {', '.join(extra_disciplines)}\n"
                        teacher_disciplines_errors.append(error)
                        teacher_disciplines_success = False

                    found = True
                    break

            if not found:
                error = f"У преподавателя {teacher} дисциплины не указаны\n"
                teacher_disciplines_errors.append(error)
                teacher_disciplines_success = False

        if teacher_disciplines_success:
            teacher_disciplines_errors.append('Дисциплины преподавателей: Успешно\n')
        teacher_disciplines_errors_str = ''.join(teacher_disciplines_errors)
        print(teacher_disciplines_errors_str)





    # Проверяем, что sved_file не равен None
    if sved_file is not None:
        tree_sved = ET.parse(os.path.join(folder_path, sved_file))
        root_sved = tree_sved.getroot()

        # Словарь для хранения дисциплин по группам из sved_file
        group_disciplines_sved = {}

        for row in root_sved.findall('row'):
            disciplines = row.find('discipline').text
            groups = row.find('group_lect').text.split()

            # Убираем "[выбранные дисциплины]" из названия дисциплины
            disciplines = disciplines.replace("[выбранные дисциплины]", "").strip()

            for group in groups:
                if group not in group_disciplines_sved:
                    group_disciplines_sved[group] = set()
                group_disciplines_sved[group].add(disciplines)


        # Сравнение дисциплин каждой группы
        group_disciplines_errors = []
        group_disciplines_success = True

        for group, disciplines in group_disciplines.items():
            if group in group_disciplines_sved:
                missing_disciplines = group_disciplines_sved[group] - disciplines
                extra_disciplines = disciplines - group_disciplines_sved[group]

                if missing_disciplines:
                    error = f"У группы {group} не хватает дисциплин: {', '.join(missing_disciplines)}\n"
                    group_disciplines_errors.append(error)
                    group_disciplines_success = False

                if extra_disciplines:
                    error = f"У группы {group} лишняя дисциплина: {', '.join(extra_disciplines)}\n"
                    group_disciplines_errors.append(error)
                    group_disciplines_success = False
            else:
                error = f"У группы {group} дисциплины не указаны\n"
                group_disciplines_errors.append(error)
                group_disciplines_success = False

        if group_disciplines_success:
            group_disciplines_errors.append('Дисциплины групп: Успешно\n')
        group_disciplines_errors_str = ''.join(group_disciplines_errors)
        print(group_disciplines_errors_str)



        # Словарь для хранения часов дисциплин по группам из sved_file
        group_disciplines_hour_sved = {}

        for row in root_sved.findall('row'):
            discipline = row.find('discipline').text
            teacher_lect = row.find('teacher_lect').text
            hour_week_lect = int(row.find('hour_week_lect').text)
            group_lect = row.find('group_lect').text.split()

            # Убираем "[выбранные дисциплины]" из названия дисциплины
            discipline = discipline.replace("[выбранные дисциплины]", "").strip()

            key = (discipline, tuple(sorted(group_lect)))
            if key not in group_disciplines_hour_sved:
                group_disciplines_hour_sved[key] = {'teacher': teacher_lect, 'hours': hour_week_lect,
                                                    'groups': set(group_lect)}
        #print(group_disciplines_hour_sved)

        #for key, info in group_disciplines_hour_sved.items():
        #    discipline, group_lect = key
        #    print(f"Дисциплина: {discipline}")
        #    print(f"Группы на лекции: {', '.join(group_lect)}")
        #    print(f"Преподаватель: {info['teacher']}")
        #    print(f"Количество часов лекций: {info['hours']}")
        #    print("----------------------")

        group_disciplines_hour_sved_new = {}

        for key, value in group_disciplines_hour_sved.items():
            discipline, groups = key
            for group in groups:
                if group not in group_disciplines_hour_sved_new:
                    group_disciplines_hour_sved_new[group] = []
                group_disciplines_hour_sved_new[group].append({
                    'discipline': discipline,
                    'teacher': value['teacher'],
                    'hours': value['hours']
                })

        #for group, disciplines in group_disciplines_hour_sved_new.items():
        #    print(f"Группа {group}:")
        #    for discipline in disciplines:
        #        print(f"  {discipline['discipline']} - {discipline['teacher']} ({discipline['hours']} ч)")
        #    print()
        #print(group_disciplines_hour_sved_new)

        disciplines_hour_lect_errors = []
        group_disciplines_hour_success = True

        for group, disciplines in final_dict.items():
            for discipline in disciplines:
                found = False
                for g, d_list in group_disciplines_hour_sved_new.items():
                    if g.strip() == group.strip():  # сравниваем строки напрямую
                        for d in d_list:
                            if d['discipline'] == discipline['discipline'] and d['teacher'].split()[0] == discipline['teacher'].split()[0]:
                                found = True
                                if discipline['total_hours'] > d['hours']:
                                    group_disciplines_hour_success = False
                                    disciplines_hour_lect_errors.append(f"У группы {group} по дисциплине {discipline['discipline']} больше часов в неделю на {discipline['total_hours'] - d['hours']}\n")
                                elif discipline['total_hours'] < d['hours']:
                                    group_disciplines_hour_success = False
                                    disciplines_hour_lect_errors.append(f"У группы {group} по дисциплине {discipline['discipline']} меньше часов в неделю на {d['hours'] - discipline['total_hours']}\n")
                                break
                if not found:
                    group_disciplines_hour_success = False
                    disciplines_hour_lect_errors.append(f"Не найдена запись для группы {group} по дисциплине {discipline['discipline']} в сведениях для составления расписания\n")

        if group_disciplines_hour_success:
            disciplines_hour_lect_errors.append('Часы в неделю по лекциям: Успешно\n')
        disciplines_hour_lect_errors_str = ''.join(disciplines_hour_lect_errors)
        print(disciplines_hour_lect_errors_str)



        # Словарь для хранения часов практик дисциплин по группам из sved_file
        group_disciplines_hour_pract = {}

        for row in root_sved.findall('.//row'):
            discipline = row.find('discipline').text.replace("[выбранные дисциплины]", "").strip()
            teacher_pract = row.find('teacher_pract').text
            group_pract = row.find('group_pract').text.replace("\n", "")
            hour_week_pract_element = row.find('hour_week_pract')
            if hour_week_pract_element.text == "None":
                continue
            hour_week_pract = int(hour_week_pract_element.text)

            if group_pract not in group_disciplines_hour_pract:
                group_disciplines_hour_pract[group_pract] = {}

            if discipline not in group_disciplines_hour_pract[group_pract]:
                group_disciplines_hour_pract[group_pract][discipline] = {}

            if teacher_pract not in group_disciplines_hour_pract[group_pract][discipline]:
                group_disciplines_hour_pract[group_pract][discipline][teacher_pract] = 0

            group_disciplines_hour_pract[group_pract][discipline][teacher_pract] += hour_week_pract

        hour_pract_success = True
        hour_pract = []
        # Проверка часов в неделю практик
        for group, disciplines in data_pract_new.items():
            for discipline, teachers in disciplines.items():
                for teacher, hours in teachers.items():
                    found = False
                    relevant_combination = None

                    # Ищем соответствующую комбинированную запись, содержащую текущую группу
                    for combined_groups, disc in group_disciplines_hour_pract.items():
                        groups = combined_groups.split()  # Разделяем строку групп
                        if group in groups:
                            relevant_combination = combined_groups
                            if discipline in disc:
                                combined_teachers = group_disciplines_hour_pract[combined_groups][discipline]
                                teacher_short = ' '.join(teacher.split()[:1])
                                combined_hours = 0

                                # Суммируем часы для всех указанных групп в data_pract_new
                                for combined_group in groups:
                                    if combined_group in data_pract_new and discipline in data_pract_new[
                                        combined_group]:
                                        if teacher in data_pract_new[combined_group][discipline]:
                                            combined_hours += data_pract_new[combined_group][discipline][teacher]

                                # Проверяем совпадение преподавателей и часов
                                for k, v in combined_teachers.items():
                                    if k.split()[0] == teacher_short:
                                        found = True
                                        if combined_hours > v:
                                            hour_pract.append(f"У группы {relevant_combination} по дисциплине {discipline} (преподаватель {k}) больше часов на {combined_hours - v}\n")
                                            hour_pract_success = False
                                        elif combined_hours < v:
                                            hour_pract.append(f"У группы {relevant_combination} по дисциплине {discipline} (преподаватель {k}) меньше часов на {v - combined_hours}\n")
                                            hour_pract_success = False
                                        break
                                break
                    if found == False and relevant_combination:
                        hour_pract.append(f"Не найдена запись для преподавателя {k} в дисциплине {discipline} для группы {relevant_combination}\n")
                        hour_pract_success = False
                    elif found == False:
                        hour_pract.append(f"Для записи в расписании группы {group} дисциплины {discipline} преподавателя {k} нет соответствующей записи в сведениях к составлению расписания\n")
                        hour_pract_success = False

        if hour_pract_success:
            hour_pract.append("Часы в неделю по практикам: Успешно\n")
        hour_pract_str = ''.join(hour_pract)
        print(hour_pract_str)

    # Создаем и записываем значения в файл report_class.xml
    report_class_root = ET.Element("report_class_rasp")
    report_class = ET.SubElement(report_class_root, "report")
    report_class.text = "fail"

    # Создаем дочерние элементы и записываем в них строки
    teacher_disciplines = ET.SubElement(report_class_root, "teacher_disciplines")
    teacher_disciplines.text = teacher_disciplines_errors_str

    group_disciplines = ET.SubElement(report_class_root, "group_disciplines")
    group_disciplines.text = group_disciplines_errors_str

    hour_lect = ET.SubElement(report_class_root, "hour_lect")
    hour_lect.text = disciplines_hour_lect_errors_str

    hour_pract = ET.SubElement(report_class_root, "hour_pract")
    hour_pract.text = hour_pract_str

    teacher_lessons_conf = ET.SubElement(report_class_root, "teacher_lessons_conf")
    teacher_lessons_conf.text = teacher_conf_str

    teacher_lessons_conf = ET.SubElement(report_class_root, "audit_capacity")
    teacher_lessons_conf.text = 'Вместимость аудиторий: Успешно '

    # Проверяем все непустые теги на наличие слова "Успешно"
    if all("Успешно" in tag.text for tag in
           [teacher_disciplines, group_disciplines, hour_lect, hour_pract, teacher_lessons_conf]):
        report_class.text = "success"

    report_class_tree = ET.ElementTree(report_class_root)
    report_class_tree.write(os.path.join(folder_path, "report_class_rasp.xml"))
    report_class_name = "report_class_rasp.xml"

    result_class = xml_to_bd(folder_path, report_class_name, folder)
    return report_class_name


# ########################################## Проверка расписания экзаменов
def check_xml_exam(output_folder, folder):
    folder_path = output_folder
    print('папка экзамена ', folder_path)
    folder = folder
    # Выбираются файлы, кроме отчетов
    files_list = [f for f in os.listdir(folder_path) if f not in ['report_class_rasp.xml', 'report_exam_rasp.xml']]

    exam_rasp_file = None
    workload_file = None
    sved_file = None

    for file in files_list:
        tree = ET.parse(os.path.join(folder_path, file))
        root = tree.getroot()
        file_type = root.find('.//file_type')
        if file_type is not None:
            if file_type.text == 'exam_rasp':
                exam_rasp_file = file
            elif file_type.text == 'workload':
                workload_file = file
            elif file_type.text == 'sved':
                sved_file = file

    tree = ET.parse(os.path.join(folder_path, exam_rasp_file))
    root = tree.getroot()

    group_exams = {}

    for exam in root.findall('.//exam'):
        group = exam.find('group').text
        start_date = exam.find('start_date').text
        day_of_week = exam.find('day_of_week').text
        start_time = exam.find('start_time').text
        key = (start_date, day_of_week, start_time)

        if group in group_exams:
            if key in group_exams[group]:
                group_exams[group][key].append(exam.find('discipline').text)
            else:
                group_exams[group][key] = [exam.find('discipline').text]
        else:
            group_exams[group] = {key: [exam.find('discipline').text]}

    exam_group_result = ""
    found_exam = False
    for group, exams in group_exams.items():
        for key, disciplines in exams.items():
            if len(disciplines) > 1:
                exam_group_result += f"У группы {group} конфликт экзаменов: {', '.join(disciplines)} ({key[0]} {key[1]} {key[2]})\n"
                found_exam = True
    if not found_exam:
        exam_group_result = "У группы одновременно один экзамен: Успешно "

    print(exam_group_result)


    # Создаем и записываем значения в файл report_exam.xml
    report_exam_root = ET.Element("report_exam_rasp")
    report_exam = ET.SubElement(report_exam_root, "report")
    report_exam.text = "fail"

    # Создаем дочерние элементы и записываем в них строки
    group_exam_conf = ET.SubElement(report_exam_root, "group_exam_conf")
    group_exam_conf.text = exam_group_result

    # Проверяем все непустые теги на наличие слова "Успешно"
    if "Успешно" in group_exam_conf.text:
        report_exam.text = "success"

    report_exam_tree = ET.ElementTree(report_exam_root)
    report_exam_tree.write(os.path.join(folder_path, "report_exam_rasp.xml"))

    report_exam_name = "report_exam_rasp.xml"

    result_exam = xml_to_bd(folder_path, report_exam_name, folder)
    return report_exam_name


# ########################################### Добавление расписания занятий в базу данных
def class_to_db(directory_path, year, semester):
    my_file = MyFiles.objects.filter(file_type='class_rasp', year=year, semester=semester).first()
    file_name_xls = my_file.file.name.rsplit('.', 1)[0]
    file_name_xml = f"{file_name_xls}.xml"
    file_path = f"{directory_path}\\{file_name_xml}"  # путь до расписания xml
    print(file_path)
#
    tree = ET.parse(file_path)
    root = tree.getroot()

    for semester_element in root.findall('.//semester'):

        # заполнение таблицы Факультет
        for faculty_element in semester_element.findall('.//faculty'):
            faculty_name = faculty_element.attrib['name']
            faculty, created = Faculty.objects.get_or_create(faculty=faculty_name)
            if created:
                print(f"Создался новый факультет: {faculty_name}")
            else:
                print(f"Факультет {faculty_name} уже существует")

            # заполнение таблицы Группа
            for course_element in faculty_element.findall('.//course'):
                for group_element in course_element.findall('.//group'):
                    group_name = group_element.attrib['name']
                    course_num = int(course_element.attrib['name'])
                    num_students_element = group_element.find('.//num_students')
                    num_students = int(num_students_element.get('text'))

                    if "ИВТМ" in group_name or "ИСТМ" in group_name:
                        qualification = Qualification.objects.get(qualification="Магистратура")
                    else:
                        qualification = Qualification.objects.get(qualification="Бакалавриат")

                    group, created = Group.objects.get_or_create(
                        group=group_name,
                        faculty=faculty,
                        qualification=qualification,
                    )

                    # обновление некоторых значений, если такая группа уже есть
                    group.course = course_num
                    group.num_students = num_students
                    group.save()

                    if created:
                        print(f"Created new group: {group_name}")
                    else:
                        print(f"Group {group_name} already exists")

                    # Collect unique num_group values for each group
                    num_groups = set()
                    for week_element in group_element.findall('week'):
                        for day_element in week_element.findall('day'):
                            for lesson_element in day_element.findall('lesson'):
                                num_group_element = lesson_element.find('num_group')
                                if num_group_element is not None:
                                    num_group = num_group_element.text
                                    if num_group is not None and num_group != '':
                                        num_groups.add(num_group)

                    # Create records in Subgroups table
                    for num_group in num_groups:
                        if num_group == 'null':
                            subgroup_number = 0
                        else:
                            subgroup_number = int(num_group)

                        subgroup, created = Subgroup.objects.get_or_create(
                            group=group,
                            subgroup_number=subgroup_number,
                            defaults={'group': group, 'subgroup_number': subgroup_number}
                        )

                        if created:
                            print(f"Created new subgroup: {subgroup_number} for group {group_name}")
                        else:
                            print(f"Subgroup {subgroup_number} for group {group_name} already exists")

                        # Collect unique discipline names from the XML file
                        disciplines = set()
                        for lesson_element in group_element.findall('.//lesson'):
                            discipline_element = lesson_element.find('.//discipline')
                            if discipline_element is not None:
                                discipline_name = discipline_element.text
                                if discipline_name is not None:
                                    disciplines.add(discipline_name)

                        # Create records in Discipline table
                        for discipline_name in disciplines:
                            discipline, created = Discipline.objects.get_or_create(discipline=discipline_name)
                            if created:
                                print(f"Created new discipline: {discipline_name}")
                            else:
                                print(f"Discipline {discipline_name} already exists")

                        # Create records in Teacher table
                        for lesson_element in group_element.findall('.//lesson'):
                            teacher_element = lesson_element.find('.//teacher')
                            if teacher_element is not None:
                                short_name = teacher_element.text
                                position_element = lesson_element.find('.//position')
                                if position_element is not None:
                                    position_name = position_element.text
                                    if position_name == 'Асс.':
                                        position_name = 'Ассистент'
                                    elif position_name == 'Доц.':
                                        position_name = 'Доцент'
                                    elif position_name == 'Проф.':
                                        position_name = 'Профессор'
                                    elif position_name == 'Ст.пр.':
                                        position_name = 'Старший преподаватель'
                                    position, created = Position.objects.get_or_create(position=position_name)
                                    if created:
                                        print(f"Created new position: {position_name}")
                                    else:
                                        print(f"Position {position_name} already exists")

                                department_element = lesson_element.find('.//department')
                                if department_element is not None:
                                    department_num = department_element.text
                                    if department_num is not None:
                                        department_number, created = Department.objects.get_or_create(
                                            department_number=department_num)
                                        if created:
                                            print(f"Created new department: {department_num}")
                                        else:
                                            print(f"Department {department_num} already exists")

                                teacher, created = Teacher.objects.get_or_create(
                                    full_name='',  # leave full_name empty
                                    short_name=short_name,
                                    position=position,
                                    department=department_number,
                                )
                                if created:
                                    print(f"Created new teacher: {short_name}")
                                else:
                                    print(f"Teacher {short_name} already exists")

                    # Create records in Lesson table
                    for week_element in group_element.findall('week'):
                        for day_element in week_element.findall('day'):
                            for lesson_element in day_element.findall('lesson'):
                                weekday_name = day_element.attrib['name']
                                weekday = Weekday.objects.get(weekday=weekday_name)

                                subgroup_numbers = []
                                for num_group_element in lesson_element.findall('num_group'):
                                    num_group_text = num_group_element.text
                                    if num_group_text == 'null':
                                        subgroup_number = 0
                                    else:
                                        subgroup_number = int(num_group_text)
                                    subgroup_numbers.append(subgroup_number)

                                start_time_text = lesson_element.find('start_time').text
                                end_time_text = lesson_element.find('end_time').text
                                start_time = dt.datetime.strptime(start_time_text, '%H:%M').time()
                                end_time = dt.datetime.strptime(end_time_text, '%H:%M').time()
                                lesson_time = LessonTime.objects.get(start_time=start_time, end_time=end_time)

                                discipline_name = lesson_element.find('discipline').text
                                discipline = Discipline.objects.get(discipline=discipline_name)

                                lesson_type_name = lesson_element.find('lesson_type').text
                                lesson_type = LessonType.objects.get(lesson_type__iexact=lesson_type_name)

                                teacher_name = lesson_element.find('teacher').text
                                teacher = Teacher.objects.get(short_name=teacher_name)

                                building_name = lesson_element.find('building').text
                                if building_name == 'Инж.корп.':
                                    center_number = 3
                                else:
                                    center_number = 3  # или любое другое значение по умолчанию

                                classroom_number = int(lesson_element.find('auditorium').text)

                                classroom, created = Classroom.objects.get_or_create(center_number=center_number,
                                                                                     classroom=classroom_number)
                                if created:
                                    classroom.num_seats = None
                                    classroom.save()

                                week_mapping = {'I': 1, 'II': 2, 'null': 0}
                                week_number = week_mapping[week_element.attrib['name']]

                                start_date_text = lesson_element.find('start_date').text
                                end_date_text = lesson_element.find('end_date').text
                                start_date_format = '%d.%m.%Y'
                                end_date_format = '%d.%m.%Y'
                                start_date_object = dt.datetime.strptime(start_date_text, start_date_format)
                                end_date_object = dt.datetime.strptime(end_date_text, end_date_format)
                                start_date = start_date_object.strftime('%Y-%m-%d')
                                end_date = end_date_object.strftime('%Y-%m-%d')

                                lesson, created = Lesson.objects.get_or_create(
                                    week_number=week_number,
                                    weekday=weekday,
                                    lesson_number=lesson_time,
                                    discipline=discipline,
                                    lesson_type=lesson_type,
                                    teacher=teacher,
                                    classroom=classroom,
                                    start_date=start_date,
                                    end_date=end_date
                                )

                                for subgroup_number in subgroup_numbers:
                                    subgroup = Subgroup.objects.get(group=group, subgroup_number=subgroup_number)
                                    lesson.subgroup.add(subgroup)  # добавляем подгруппу к занятию
                                lesson.save()

    print('class')


# ########################################### Добавление расписания экзаменов в базу данных
def exam_to_db(directory_path, year, semester):
    my_file = MyFiles.objects.filter(file_type='exam_rasp', year=year, semester=semester).first()
    file_name_xls = my_file.file.name.rsplit('.', 1)[0]
    file_name_xml = f"{file_name_xls}.xml"
    file_path = f"{directory_path}\\{file_name_xml}"  # путь до расписания xml
    print(file_path)

    tree = ET.parse(file_path)
    root = tree.getroot()

    for exam in root.findall('exam'):
        semester_index = exam.find('semester_index').text
        faculty = exam.find('faculty').text
        course = exam.find('course').text
        group = exam.find('group').text
        discipline = exam.find('discipline').text
        department = exam.find('department').text
        position = exam.find('position').text
        teacher = exam.find('teacher').text
        num_students = exam.find('num_students').text
        lesson_type = exam.find('lesson_type').text
        building = exam.find('building').text
        start_date_str = exam.find('start_date').text
        day_of_week = exam.find('day_of_week').text
        start_time = exam.find('start_time').text

        # Find the group
        faculty = Faculty.objects.filter(faculty=faculty).first()
        if not faculty:
            print(f"Faculty '{faculty}' not found")
            continue
        group = Group.objects.filter(group=group, faculty=faculty, course=course).first()
        if not group:
            print(f"Group '{group}' not found")
            continue

        # Find the discipline
        discipline = Discipline.objects.filter(discipline=discipline).first()
        if not discipline:
            print(f"Discipline '{discipline}' not found")
            continue

        # Find the teacher
        teacher = Teacher.objects.filter(short_name=teacher).first()
        if not teacher:
            print(f"Teacher '{teacher}' not found")
            continue

        # Find the lesson
        lesson_type = LessonType.objects.filter(lesson_type="Лекция").first()
        if not lesson_type:
            print(f"Lesson type '{lesson_type}' not found")
            continue

        exam_time = ExamTime.objects.get_or_create(start_time=start_time)[0]

        start_date_obj = dt.datetime.strptime(start_date_str, '%d.%m.%Y')
        start_date_db = start_date_obj.strftime('%Y-%m-%d')

        print(f"Creating exam for lesson: {discipline}, {teacher}, {group}, {lesson_type}")

        lesson = Lesson.objects.filter(discipline=discipline, teacher=teacher, subgroup__group=group,
                                       lesson_type=lesson_type).first()
        if not lesson:
            print(
                f"Lesson not found for discipline '{discipline}', teacher '{teacher}', group '{group}' and lesson type '{lesson_type}'")
            continue

        exam, created = Exam.objects.get_or_create(lesson=lesson, exam_date=start_date_db, start_time=exam_time,
                                                   defaults={'exam_date': start_date_db, 'start_time': exam_time})
        if not created:
            exam.exam_date = start_date_db
            exam.start_time = exam_time
            exam.save()

    print('Exams added to database')
